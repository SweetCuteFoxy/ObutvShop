import openpyxl
import psycopg2
from datetime import datetime

conn = psycopg2.connect(
    host="localhost", port=5432,
    dbname="library", user="postgres", password="postgres"
)
conn.autocommit = True
cur = conn.cursor()

# === Создание таблиц ===

cur.execute("""
DROP TABLE IF EXISTS book_loans CASCADE;
DROP TABLE IF EXISTS books CASCADE;
DROP TABLE IF EXISTS users CASCADE;
DROP TABLE IF EXISTS genres CASCADE;
DROP TABLE IF EXISTS publishers CASCADE;
DROP TABLE IF EXISTS roles CASCADE;
DROP TABLE IF EXISTS loan_statuses CASCADE;
""")

cur.execute("""
CREATE TABLE roles (
    id SERIAL PRIMARY KEY,
    name VARCHAR(50) NOT NULL UNIQUE
);

CREATE TABLE genres (
    id SERIAL PRIMARY KEY,
    name VARCHAR(100) NOT NULL UNIQUE
);

CREATE TABLE publishers (
    id SERIAL PRIMARY KEY,
    name VARCHAR(200) NOT NULL UNIQUE
);

CREATE TABLE loan_statuses (
    id SERIAL PRIMARY KEY,
    name VARCHAR(50) NOT NULL UNIQUE
);

CREATE TABLE users (
    id SERIAL PRIMARY KEY,
    full_name VARCHAR(200) NOT NULL,
    library_card VARCHAR(20) NOT NULL UNIQUE,
    login VARCHAR(100) NOT NULL UNIQUE,
    password_text VARCHAR(100) NOT NULL,
    role_id INTEGER NOT NULL REFERENCES roles(id)
);

CREATE TABLE books (
    id SERIAL PRIMARY KEY,
    isbn VARCHAR(30) NOT NULL UNIQUE,
    title VARCHAR(300) NOT NULL,
    author VARCHAR(200) NOT NULL,
    genre_id INTEGER NOT NULL REFERENCES genres(id),
    publisher_id INTEGER NOT NULL REFERENCES publishers(id),
    year_published INTEGER NOT NULL,
    pages INTEGER NOT NULL,
    total_copies INTEGER NOT NULL DEFAULT 0,
    available_copies INTEGER NOT NULL DEFAULT 0,
    annotation TEXT,
    cover_image VARCHAR(200)
);

CREATE TABLE book_loans (
    id SERIAL PRIMARY KEY,
    user_id INTEGER NOT NULL REFERENCES users(id),
    book_id INTEGER NOT NULL REFERENCES books(id),
    loan_date DATE NOT NULL,
    return_date_expected DATE NOT NULL,
    return_date_actual DATE,
    status_id INTEGER NOT NULL REFERENCES loan_statuses(id)
);
""")

print("Таблицы созданы")

# === Загрузка данных из файлов импорта ===

base = r"C:\Users\edikk\Desktop\DEMO_E\Library\Прил_2_ОЗ_КОД 09.02.07-2-2026-М1\import"

# -- Роли (из roles_import.xlsx) --
roles_map = {}
wb_roles = openpyxl.load_workbook(f"{base}\\roles_import.xlsx")
ws_roles = wb_roles.active
for row in ws_roles.iter_rows(min_row=2, values_only=True):
    rname = row[0]
    if not rname:
        continue
    cur.execute("INSERT INTO roles (name) VALUES (%s) RETURNING id", (rname,))
    roles_map[rname] = cur.fetchone()[0]
print(f"Ролей: {len(roles_map)}")

# -- Статусы выдачи (из loan_statuses_import.xlsx) --
statuses_map = {}
wb_st = openpyxl.load_workbook(f"{base}\\loan_statuses_import.xlsx")
ws_st = wb_st.active
for row in ws_st.iter_rows(min_row=2, values_only=True):
    sname = row[0]
    if not sname:
        continue
    cur.execute("INSERT INTO loan_statuses (name) VALUES (%s) RETURNING id", (sname,))
    statuses_map[sname] = cur.fetchone()[0]
print(f"Статусов: {len(statuses_map)}")

# -- Жанры (из genres_import.xlsx) --
genres_map = {}
wb_g = openpyxl.load_workbook(f"{base}\\genres_import.xlsx")
ws_g = wb_g.active
for row in ws_g.iter_rows(min_row=2, values_only=True):
    gname = row[0]
    if not gname:
        continue
    cur.execute("INSERT INTO genres (name) VALUES (%s) RETURNING id", (gname,))
    genres_map[gname] = cur.fetchone()[0]
print(f"Жанров: {len(genres_map)}")

# -- Издательства (из publishers_import.xlsx) --
publishers_map = {}
wb_p = openpyxl.load_workbook(f"{base}\\publishers_import.xlsx")
ws_p = wb_p.active
for row in ws_p.iter_rows(min_row=2, values_only=True):
    pname = row[0]
    if not pname:
        continue
    cur.execute("INSERT INTO publishers (name) VALUES (%s) RETURNING id", (pname,))
    publishers_map[pname] = cur.fetchone()[0]
print(f"Издательств: {len(publishers_map)}")

# -- Пользователи --
wb = openpyxl.load_workbook(f"{base}\\library_users_import.xlsx")
ws = wb.active
users_map = {}  # library_card -> user_id
for row in ws.iter_rows(min_row=2, values_only=True):
    role, full_name, card, login, pwd = row[0], row[1], row[2], row[3], row[4]
    if not role or not full_name:
        continue
    rid = roles_map.get(role)
    if rid is None:
        continue
    cur.execute(
        "INSERT INTO users (full_name, library_card, login, password_text, role_id) "
        "VALUES (%s, %s, %s, %s, %s) RETURNING id",
        (full_name, card, login, pwd, rid)
    )
    users_map[card] = cur.fetchone()[0]

print(f"Пользователей: {len(users_map)}")

# -- Книги (из books_import.xlsx) --
wb2 = openpyxl.load_workbook(f"{base}\\books_import.xlsx")
ws2 = wb2.active
books_map = {}  # isbn -> book_id

for row in ws2.iter_rows(min_row=2, values_only=True):
    isbn, title, author, genre, publisher, year, pages, total, avail, annot = row[:10]
    if not isbn or not title:
        continue

    isbn = str(isbn).strip()
    title = str(title).strip()

    gid = genres_map.get(genre)
    pid = publishers_map.get(publisher)
    if gid is None or pid is None:
        print(f"  skip book {isbn}: genre={genre} publisher={publisher}")
        continue

    cur.execute(
        "INSERT INTO books (isbn, title, author, genre_id, publisher_id, year_published, pages, total_copies, available_copies, annotation) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id",
        (isbn, title, author, gid, pid,
         int(year), int(pages), int(total), int(avail), annot)
    )
    books_map[isbn] = cur.fetchone()[0]

print(f"Книг: {len(books_map)}")

# -- Выдачи --
wb3 = openpyxl.load_workbook(f"{base}\\book_loans_import.xlsx")
ws3 = wb3.active
loan_count = 0
for row in ws3.iter_rows(min_row=2, values_only=True):
    lid, card, isbn, loan_dt, ret_exp, ret_act, status = row[:7]
    if not lid or not card:
        continue

    uid = users_map.get(card)
    bid = books_map.get(str(isbn).strip())
    if uid is None or bid is None:
        print(f"  skip loan {lid}: card={card} isbn={isbn}")
        continue

    sid = statuses_map.get(status, statuses_map["На руках"])

    ret_actual = None
    if ret_act and ret_act != '-' and isinstance(ret_act, datetime):
        ret_actual = ret_act.date()

    cur.execute(
        "INSERT INTO book_loans (user_id, book_id, loan_date, return_date_expected, return_date_actual, status_id) "
        "VALUES (%s, %s, %s, %s, %s, %s)",
        (uid, bid, loan_dt.date() if isinstance(loan_dt, datetime) else loan_dt,
         ret_exp.date() if isinstance(ret_exp, datetime) else ret_exp,
         ret_actual, sid)
    )
    loan_count += 1

print(f"Выдач: {loan_count}")

cur.close()
conn.close()
print("Готово!")
